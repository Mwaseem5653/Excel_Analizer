import pandas as pd
import os
import re
import requests
import base64
import concurrent.futures
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from utils.table_header_finder import read_excel_auto
from utils.api_clients import get_phone_info, get_eyecon_info


def analyze_excel(file_path, top_n=15, enable_lookup=True, enable_eyecon_lookup=False, eyecon_top_n=15, include_eyecon_images=False):

    # -------------------- Read Excel --------------------
    file_ext = os.path.splitext(file_path)[1].lower()
    if file_ext == ".csv":
        df = pd.read_csv(file_path)
    else:
        df = read_excel_auto(file_path)

    df_original = df.copy()  # FORMATTED RAW BACKUP
    
    # -------------------- Column Detection --------------------
    def find_col(possible):
        return next((c for c in df.columns if c.strip().lower() in
                     [x.lower() for x in possible]), None)

    b_col = find_col(["B Number", "BNUMBER", "b number", "b party", "b_party", "CALL_DIALED_NUM", "BParty"])
    if not b_col:
        raise ValueError("❌ B Party column not found")

    calltype_col = find_col([
        "CallType", "CALL_TYPE", "Type"
    ])

    

    date_col = find_col([
        "CALL_START_DT_TM", "Start Date", "Datetime", "Date", "STRT_TM","Start Time"
    ])

    inbound = find_col([
        "INBOUND_OUTBOUND_IND","Direction"
    ])

    address_col = find_col(["Address", "Location", "Addr", "SITE_ADDRESS", "SiteLocation"])
    imei_col = find_col(["IMEI", "imei", "Imei number", "IMEI numbe"])

    # -------------------- Date --------------------
    if date_col:
        df["__DATE__"] = pd.to_datetime(df[date_col], errors="coerce")
    else:
        df["__DATE__"] = None

    # -------------------- PREPARE FORMATTED SHEET --------------------
    formatted_df = df_original.copy()
    
    # 1. Format Date Column to String (if exists)
    if date_col and date_col in formatted_df.columns:
        # Check if it's already datetime, if not convert
        if not pd.api.types.is_datetime64_any_dtype(formatted_df[date_col]):
             formatted_df[date_col] = pd.to_datetime(formatted_df[date_col], errors='coerce')
        
        # Format as string
        formatted_df[date_col] = formatted_df[date_col].dt.strftime('%Y-%m-%d %H:%M:%S').fillna('')

    # 2. Format Numeric Columns (Skip Date)
    for col in formatted_df.columns:
        if col == date_col:
            continue
            
        if pd.api.types.is_numeric_dtype(formatted_df[col]):
            def format_value(x):
                if pd.isna(x):
                    return x
                # If it has decimals, keep them
                if isinstance(x, float) and not x.is_integer():
                    return str(x)
                # Otherwise format as int string with leading space (for Excel text format)
                return f" {int(x)}"

            formatted_df[col] = formatted_df[col].apply(format_value)

    # -------------------- CLEAN NUMBER (ANALYSIS ONLY) --------------------
    def normalize(num):
        if pd.isna(num):
            return None
        num = re.sub(r"\D", "", str(num))
        if num.startswith("92"):
            num = num[2:]
        elif num.startswith("0"):
            num = num[1:]
        return num if re.fullmatch(r"3\d{9}", num) else None

    df["__B_CLEAN__"] = df[b_col].apply(normalize)

    # -------------------- MOBILE SUMMARY --------------------
    mob = df.dropna(subset=["__B_CLEAN__"])
    g = mob.groupby("__B_CLEAN__")

    mobile_summary = pd.DataFrame({
        "Mobile Number": g.size().index,
        "Starting Date": g["__DATE__"].min().values,
        "Ending Date": g["__DATE__"].max().values,
        "Count": g.size().values
    }).sort_values("Count", ascending=False)

    # -------------------- FETCH API INFO (PARALLEL) --------------------
    lookup_cache = {} # Store results to use in Call Logs later

    if enable_lookup:
        mobile_summary["Name"] = ""
        mobile_summary["CNIC"] = ""
        mobile_summary["Address"] = ""

        # Prepare list for parallel execution
        top_indices = mobile_summary.index[:top_n]
        
        def fetch_phone_details(idx):
            raw_num = str(mobile_summary.at[idx, "Mobile Number"]).strip()
            query_num = "0" + raw_num
            try:
                data = get_phone_info(query_num)
                if isinstance(data, list) and len(data) > 0:
                    record = data[0]
                    name_val = record.get("name") or record.get("Name") or ""
                    cnic_val = record.get("cnic") or record.get("CNIC") or ""
                    addr_val = record.get("address") or record.get("Address") or ""
                    return idx, raw_num, name_val, cnic_val, addr_val
            except Exception:
                pass
            return idx, raw_num, None, None, None

        # Execute sequentially (Single Thread)
        for idx in top_indices:
            idx, raw_num, name_val, cnic_val, addr_val = fetch_phone_details(idx)
            
            if name_val is not None:
                mobile_summary.at[idx, "Name"] = name_val
                mobile_summary.at[idx, "CNIC"] = " " + str(cnic_val) if cnic_val else ""
                mobile_summary.at[idx, "Address"] = addr_val
                
                if raw_num not in lookup_cache:
                        lookup_cache[raw_num] = {}
                lookup_cache[raw_num]["Name"] = name_val
                lookup_cache[raw_num]["CNIC"] = " " + str(cnic_val) if cnic_val else ""
                lookup_cache[raw_num]["Address"] = addr_val

    # -------------------- FETCH EYECON INFO (PARALLEL) --------------------
    eyecon_cache = {}
    
    # Helper to extract image URL (runs inside thread)
    def process_eyecon_image_url(data):
        image_url = ""
        
        # Helper to get image from 'images' list (usually HTTP)
        def get_images_url(d):
            if not isinstance(d, dict): return None
            if isinstance(d.get("images"), list):
                for img_entry in d["images"]:
                    if isinstance(img_entry, dict) and "pictures" in img_entry:
                        pics = img_entry["pictures"]
                        if isinstance(pics, dict):
                            return pics.get("200") or pics.get("600") or next(iter(pics.values()), None)
            return None

        # 1. PRIORITIZE HTTP URL (Nested or direct)
        if isinstance(data.get("data"), dict) and data["data"].get("photo"):
            image_url = data["data"].get("photo")
        elif data.get("photo"):
            image_url = data.get("photo")
        
        if not image_url or not str(image_url).startswith("http"):
            image_url = get_images_url(data) or get_images_url(data.get("data"))

        # 2. FALLBACK TO BASE64 if no HTTP URL found
        if not image_url:
            raw_b64 = data.get("b64") or (data.get("data", {}).get("b64") if isinstance(data.get("data"), dict) else None)
            if raw_b64:
                image_url = raw_b64 if str(raw_b64).startswith("data:image") else f"data:image/jpeg;base64,{raw_b64}"
        
        return image_url

    def extract_eyecon_names(d):
        found = set()
        if not isinstance(d, dict): return found
        fname = d.get("fullName") or d.get("name")
        if fname: found.add(fname)
        others = d.get("otherNames", [])
        if isinstance(others, list):
            for o in others:
                if isinstance(o, dict):
                    n = o.get("name")
                    if n: found.add(n)
                elif isinstance(o, str):
                    found.add(o)
        return found

    if enable_eyecon_lookup:
        mobile_summary["Eyecon Name"] = ""
        if include_eyecon_images:
            mobile_summary["Eyecon Image"] = ""
            mobile_summary["Facebook Link"] = ""
        
        top_eyecon_indices = mobile_summary.index[:eyecon_top_n]

        def fetch_eyecon_details(idx):
            raw_num = str(mobile_summary.at[idx, "Mobile Number"]).strip()
            try:
                data = get_eyecon_info(raw_num)
                
                if isinstance(data, dict) and "message" in data:
                    msg = data["message"].lower()
                    if "quota" in msg or "subscribe" in msg:
                        return idx, raw_num, "Quota Exceeded", "", ""
                
                if isinstance(data, dict) and (data.get("status") or data.get("fullName")):
                    all_names = extract_eyecon_names(data)
                    if isinstance(data.get("data"), dict):
                        all_names.update(extract_eyecon_names(data["data"]))
                    
                    final_name = " | ".join(sorted(list(all_names))) if all_names else ""
                    image_url = ""
                    fb_url = ""
                    
                    if include_eyecon_images:
                        image_url = process_eyecon_image_url(data)
                        if isinstance(data.get("facebookID"), dict):
                            fb_url = data["facebookID"].get("url")
                        elif isinstance(data.get("data"), dict) and isinstance(data["data"].get("facebookID"), dict):
                            fb_url = data["data"]["facebookID"].get("url")
                    
                    return idx, raw_num, final_name, image_url, fb_url
                
            except Exception:
                pass
            
            return idx, raw_num, "", "", ""

        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            future_to_idx = {executor.submit(fetch_eyecon_details, idx): idx for idx in top_eyecon_indices}
            
            for future in concurrent.futures.as_completed(future_to_idx):
                idx, raw_num, name, img, fb = future.result()
                if name or img:
                    mobile_summary.at[idx, "Eyecon Name"] = name
                    if include_eyecon_images:
                        mobile_summary.at[idx, "Eyecon Image"] = img
                        mobile_summary.at[idx, "Facebook Link"] = fb
                    
                    eyecon_cache[raw_num] = {"name": name, "image": img, "fb_url": fb}

    # Reorder columns
    base_mob_cols = ["Mobile Number"]
    if enable_eyecon_lookup:
        base_mob_cols.append("Eyecon Name")
        if include_eyecon_images:
            base_mob_cols.append("Eyecon Image")
            base_mob_cols.append("Facebook Link")
    if enable_lookup:
        base_mob_cols.extend(["Name", "CNIC", "Address"])
        
    desired_order = base_mob_cols + ["Starting Date", "Ending Date", "Count"] 
    final_order = [c for c in desired_order if c in mobile_summary.columns]
    mobile_summary = mobile_summary[final_order]

    # -------------------- ADDRESS & IMEI SUMMARY --------------------
    address_summary = None
    if address_col:
        g = df.groupby(address_col)
        address_summary = pd.DataFrame({
            address_col: g.size().index,
            "Starting Date": g["__DATE__"].min().values,
            "Ending Date": g["__DATE__"].max().values,
            "Count": g.size().values
        }).sort_values("Count", ascending=False)

    imei_summary = None
    if imei_col:
        g = df.groupby(imei_col)
        imei_summary = pd.DataFrame({
            "IMEI Number": g.size().index,
            "Starting Date": g["__DATE__"].min().values,
            "Ending Date": g["__DATE__"].max().values,
            "Count": g.size().values
        }).sort_values("Count", ascending=False)
    if imei_summary is not None:
        imei_summary["IMEI Number"] = imei_summary["IMEI Number"].astype(str).apply(lambda x: " " + x)


    # -------------------- CALL LOGS SHEET --------------------
    call_df = df_original.copy()
    call_df["__B_CLEAN__"] = df["__B_CLEAN__"]
    call_df["__DATE__"] = df["__DATE__"]

    if inbound and calltype_col:
            call_df["CALLTYPE"] = (call_df[inbound] + " " + call_df[calltype_col])
    else:
            call_df["CALLTYPE"] = call_df[calltype_col]

    call_df["CALLTYPE"] = (call_df["CALLTYPE"].str.lower().str.replace("-", " ", regex=False).str.replace(r"\s+", " ", regex=True).str.strip())

    def is_in_call(x): return x in ["incoming", "incoming call", "incomingcall","call incoming", "callincomig","voice incoming","voiceincoming","incoming voice","incomingvoice"]
    def is_out_call(x): return x in ["outgoing", "outgoing call", "outgoingcall","call outgoing", "calloutgoing","voice outgoing","voiceoutgoing","outgoing voice","outgoingvoice"]
    def is_in_sms(x): return x in ["incoming sms","incomingsms","sms incoming", "smsincoming"]
    def is_out_sms(x): return x in ["outgoing sms", "outgoingsms","sms outgoing", "smsoutgoing"]

    summary = (
        call_df.dropna(subset=["__B_CLEAN__"]).groupby("__B_CLEAN__")
        .apply(lambda x: pd.Series({
            "Starting Date": x["__DATE__"].min(),
            "Ending Date": x["__DATE__"].max(),
            "In-SMS": x["CALLTYPE"].apply(is_in_sms).sum(),
            "Out-SMS": x["CALLTYPE"].apply(is_out_sms).sum(),
            "In-Call": x["CALLTYPE"].apply(is_in_call).sum(),
            "Out-Call": x["CALLTYPE"].apply(is_out_call).sum(),
            "Same-Num-Count": len(x),
        }))
        .reset_index().rename(columns={"__B_CLEAN__": "B-party"})
        .sort_values(by="Same-Num-Count", ascending=False).reset_index(drop=True)
    )

    base_cols = ["B-party"]
    if enable_eyecon_lookup:
        summary["Eyecon Name"] = summary["B-party"].map(lambda x: eyecon_cache.get(str(x), {}).get("name", ""))
        base_cols.append("Eyecon Name")
        if include_eyecon_images:
            summary["Eyecon Image"] = summary["B-party"].map(lambda x: eyecon_cache.get(str(x), {}).get("image", ""))
            summary["Facebook Link"] = summary["B-party"].map(lambda x: eyecon_cache.get(str(x), {}).get("fb_url", ""))
            base_cols.extend(["Eyecon Image", "Facebook Link"])
        
    if enable_lookup:
        summary["Name"] = summary["B-party"].map(lambda x: lookup_cache.get(str(x), {}).get("Name", ""))
        summary["CNIC"] = summary["B-party"].map(lambda x: lookup_cache.get(str(x), {}).get("CNIC", ""))
        summary["Address"] = summary["B-party"].map(lambda x: lookup_cache.get(str(x), {}).get("Address", ""))
        base_cols.extend(["Name", "CNIC", "Address"])

    sum_cols = list(summary.columns)
    other_cols = [c for c in sum_cols if c not in base_cols]
    summary = summary[base_cols + other_cols]

    # -------------------- SAVE --------------------
    out_dir = "temp_uploads"
    os.makedirs(out_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    out_path = os.path.join(out_dir, f"{base_name}_analyzed.xlsx")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        mobile_summary.to_excel(writer, sheet_name="Mobile Numbers", index=False)
        if address_summary is not None:
            address_summary.to_excel(writer, sheet_name="Addresses", index=False)
        if imei_summary is not None:
            imei_summary.to_excel(writer, sheet_name="IMEI Numbers", index=False)
        summary.to_excel(writer, sheet_name="Call Logs", index=False)
        formatted_df.to_excel(writer, sheet_name="Formatted Data", index=False)

    # -------------------- FORMAT --------------------
    wb = load_workbook(out_path)
    fill = PatternFill("solid", start_color="ADD8E6")
    bold = Font(bold=True)

    for ws in wb.worksheets:
        eyecon_img_idx = None
        fb_link_idx = None
        eyecon_name_idx = None
        
        # 1. Header Styling and Column Identification
        for c in ws[1]:
            c.fill = fill
            c.font = bold
            c.alignment = Alignment(horizontal="center")
            val = str(c.value) if c.value else ""
            if val == "Eyecon Image": eyecon_img_idx = c.column
            if val == "Facebook Link": fb_link_idx = c.column
            if val == "Eyecon Name": eyecon_name_idx = c.column

        # 2. Global Alignment and Width
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            # Default width
            max_len = 10
            for cell in col:
                if cell.row == 1: continue # Skip header for width calc if needed, or include
                # Apply default center alignment
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_len: max_len = length
            
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50) # Cap width

        # 3. Specific Formatting for Eyecon Name (Wrapping)
        if eyecon_name_idx:
            col_letter = get_column_letter(eyecon_name_idx)
            ws.column_dimensions[col_letter].width = 30
            for cell in ws[col_letter]:
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Fix width for link columns (override global large width)
        if eyecon_img_idx:
            ws.column_dimensions[get_column_letter(eyecon_img_idx)].width = 15
        if fb_link_idx:
            ws.column_dimensions[get_column_letter(fb_link_idx)].width = 15

        # 4. Specific Formatting for Hyperlinks
        def apply_hyperlinks(col_idx, label):
            if not col_idx: return
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                url = str(cell.value).strip() if cell.value else ""
                
                if url:
                    cell.value = label
                    # Excel has a limit of ~32k characters for hyperlinks. 
                    # If it's a normal URL it works, if it's base64 it might be too long.
                    if len(url) < 30000:
                        cell.hyperlink = url
                        cell.font = Font(color="0000FF", underline="single")
                    else:
                        cell.value = label + " (Data too long)"
                else:
                    cell.value = ""

        apply_hyperlinks(eyecon_img_idx, "View Image")
        apply_hyperlinks(fb_link_idx, "View Profile")

    if "Formatted Data" in wb.sheetnames:
        ws = wb["Formatted Data"]
        for col in ws.columns:
            for cell in col:
                cell.number_format = "@"

    wb.save(out_path)
    return out_path
